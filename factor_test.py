import os
from glob import glob

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from tqdm import tqdm

from setting import freq, freq_dict, criteria_dict, criteria_operator, econ_criteria_dict, factor_mask


def equity(data):
    """
    新增權益曲線/高點&DD欄位
    """
    dataframe = data.copy()
    dataframe['權益'] = dataframe['獲利(¤)'].sort_index().cumsum()
    dataframe['權益高點'] = dataframe['權益'].sort_index().cummax()
    dataframe['DD'] = dataframe['權益高點'] - dataframe['權益']
    dataframe['DD_pct'] = dataframe['DD'] / dataframe['權益高點'].apply(lambda x: 1 if x <= 0 else x)
    dataframe['DD_pct'] = dataframe['DD_pct']
    return dataframe


def freq_title(sheet, col):
    """
    新增header
    :param sheet: openpyxl's ws
    :param col: current column
    :return: updated current column
    """
    init_color = 13434879  # 0xccffff
    for tmp in ['月', '季', '半年']:
        sheet.cell(row=3, column=col, value=tmp).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=3, column=col).fill = PatternFill("solid", fgColor=hex(init_color).split('x')[-1])
        sheet.cell(row=4, column=col, value='績效').alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=4, column=col).border = Border(left=thin)
        sheet.cell(row=4, column=col + 1, value='MDD').alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=4, column=col + 1).border = Border(right=thin)
        sheet.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        sheet.cell(row=3, column=col).border = Border(left=thin, right=thin)
        col += 2
        init_color -= 3355392
    return col


def fill(r, col, item, sheet):
    """
    :param r: row
    :param col: col
    :param item: Pnl/MDD
    :param sheet: openpyxl's ws
    """
    if item == 100:  # for (MDD==0), 但可能有例外
        color = "808080"
    elif item > 50:
        color = "FF6600"
    elif item > 25:
        color = "FFCC00"
    elif item > 0:
        color = "C0C0C0"
    else:
        color = "808080"
    sheet.cell(row=r, column=col).fill = PatternFill("solid", fgColor=color)


def description(r, sheet):
    """
    :param r: row(工作表最底部)
    :param sheet: openpyxl's ws
    :return:
    """
    sheet.cell(row=r, column=2, value='>50').fill = PatternFill("solid", fgColor='FF6600')
    sheet.cell(row=r, column=2).border = Border(top=thin, left=thin, right=thin)
    sheet.cell(row=r + 1, column=2, value='>25').fill = PatternFill("solid", fgColor='FFCC00')
    sheet.cell(row=r + 1, column=2).border = Border(left=thin, right=thin)
    sheet.cell(row=r + 2, column=2, value='>0').fill = PatternFill("solid", fgColor='C0C0C0')
    sheet.cell(row=r + 2, column=2).border = Border(left=thin, right=thin)
    sheet.cell(row=r + 3, column=2, value='<=0').fill = PatternFill("solid", fgColor='808080')
    sheet.cell(row=r + 3, column=2).border = Border(left=thin, right=thin, bottom=thin)


def merge_category(cat, sheet):
    r = 5
    for ca in cat:
        # 避免讀到暫存檔
        ca_len = len(set(glob(f'StrategyReport-1/{ca}/*/*.xls')) ^ set(glob(f'StrategyReport-1/{ca}/*/~*.xls')))
        # 合併儲存格
        sheet.merge_cells(start_row=r, start_column=1, end_row=r + ca_len - 1, end_column=1)
        sheet.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        r += ca_len


def get_file():
    """
    :return: 報表檔名, 商品分類
    """
    # 取得所有報表檔名
    fi = glob('StrategyReport-1/*/*/*.xls*')
    # 避免讀到暫存檔
    for _ in glob('StrategyReport-1/*/*/~*.xls*'):
        fi.remove(_)
    ca = list(set(_.split('\\')[1] for _ in fi))
    ca.sort()
    return fi, ca


def get_factor():
    """
    :return: factor list
    """
    fa = glob('factor/report/*.xlsx')
    fa = [_.split('\\')[-1].replace(".xlsx", "") for _ in fa]
    return fa


def strategy_title(sheet):
    """
    策略名稱欄位
    :param sheet: openpyxl's ws
    """
    sheet.cell(row=4, column=2).border = Border(bottom=thin)
    sheet.cell(row=4, column=1).border = Border(bottom=thin)
    sheet.merge_cells(start_row=1, start_column=1, end_row=4, end_column=2)
    sheet.column_dimensions["B"].bestFit = True


def strategy_n_category(workbook, sheet, ca, stra, r):
    """
    寫入策略名稱/類別
    :param workbook: openpyxl's workbook
    :param sheet: sheet name
    :param ca: category
    :param stra: strategy name
    :param r: row
    """
    workbook[sheet].cell(row=r, column=1, value=ca)
    workbook[sheet].cell(row=r, column=2, value=stra)
    workbook[sheet].cell(row=r, column=2).border = Border(right=thin)


if __name__ == "__main__":
    # excel框線
    thin = Side(border_style="thin", color="000000")

    # 取得報表檔名&類別
    file, category = get_file()
    factor = get_factor()
    # 生成篩後結果excel
    wb_all = Workbook()
    ws = wb_all.active

    for cr in criteria_dict:
        # 建立工作表
        wb_all.create_sheet(cr)
        # 策略名稱欄位
        strategy_title(wb_all[cr])
        # 績效Factor
        wb_all[cr].cell(row=1, column=3, value=cr)
        space = 3 * 2 * len(criteria_dict[cr])
        wb_all[cr].merge_cells(start_row=1, start_column=3, end_row=1, end_column=3 + space - 1)
        wb_all[cr].cell(row=1, column=3).alignment = Alignment(horizontal="center", vertical="center")
        wb_all[cr].cell(row=1, column=3).border = Border(left=thin, right=thin)
        column = 3
        # Factor's criteria
        for i in criteria_dict[cr]:
            wb_all[cr].cell(row=2, column=column, value=criteria_operator[cr] + str(i))
            wb_all[cr].cell(row=2, column=column).alignment = Alignment(horizontal="center", vertical="center")
            wb_all[cr].merge_cells(start_row=2, start_column=column, end_row=2, end_column=column + 5)
            wb_all[cr].cell(row=2, column=column).border = Border(left=thin, right=thin)
            # freq
            column = freq_title(wb_all[cr], column)

    factor_criteria_num = 0
    for _ in econ_criteria_dict:
        factor_criteria_num += len(econ_criteria_dict[_])

    for _ in factor:
        if _ not in factor_mask:
            wb_all.create_sheet(_)
            # 策略名稱欄位
            strategy_title(wb_all[_])
            column = 3
            for ec in econ_criteria_dict:
                wb_all[_].cell(row=1, column=column, value=ec)
                space = 3 * 2 * len(econ_criteria_dict[ec])
                wb_all[_].merge_cells(start_row=1, start_column=column, end_row=1, end_column=column + space - 1)
                wb_all[_].cell(row=1, column=column).alignment = Alignment(horizontal="center", vertical="center")
                wb_all[_].cell(row=1, column=column).border = Border(left=thin, right=thin)
                for i in econ_criteria_dict[ec]:
                    wb_all[_].cell(row=2, column=column, value=criteria_operator[ec] + str(i))
                    wb_all[_].cell(row=2, column=column).alignment = Alignment(horizontal="center", vertical="center")
                    wb_all[_].merge_cells(start_row=2, start_column=column, end_row=2, end_column=column + 5)
                    wb_all[_].cell(row=2, column=column).border = Border(left=thin, right=thin)
                    # freq
                    column = freq_title(wb_all[_], column)
        else:
            continue
    # 刪除預設空白工作表
    wb_all.remove(ws)

    row = 5
    for f in tqdm(file):
        column = 3
        # 策略名稱
        strategy_name = f.split(" ")[2].replace(";", "")
        # 策略分類
        cat_name = f.split('\\')[1]

        for cr in criteria_dict:
            strategy_n_category(wb_all, cr, cat_name, strategy_name, row)
        for fa in factor:
            if fa not in factor_mask:
                strategy_n_category(wb_all, fa, cat_name, strategy_name, row)
            else:
                continue

        # 檢查資料夾
        if not os.path.exists('period_report'):
            os.makedirs('period_report')
        if not os.path.exists(f"period_report/{cat_name}"):
            os.makedirs(f"period_report/{cat_name}")
        # 建立空白xlsx
        wb = Workbook()
        wb.save(f"period_report/{cat_name}/{strategy_name}.xlsx")

        # 切出日報酬(持倉損益)
        trade = pd.read_excel(f, sheet_name='週期性分析')
        end = trade[trade['Daily Period Analysis'].isin(['Daily Rolling Period Analysis'])].index.values[0]
        daily = trade[0:end]
        daily = daily[daily['Daily Period Analysis'].notna()].reset_index(drop=True)
        daily = daily.set_axis(list(daily.iloc[0]), axis=1, inplace=False)[1:].reset_index(drop=True)
        # 確保欄位名稱統一
        daily = daily.rename(columns={daily.columns[1]: '獲利(¤)',
                                      daily.columns[-1]: '勝率',
                                      '毛利': 'pnl_w',
                                      '毛損': 'pnl_l',
                                      '交易次數': '數量'})
        # 轉換為可計算型別
        daily = daily.astype({'獲利(¤)': 'int',
                              '獲利(%)': 'float',
                              '數量': 'int'})
        daily['期間'] = pd.to_datetime(daily['期間'])
        # 以日期為index
        daily = daily.set_index('期間')
        # 添補缺失日期
        idx = pd.date_range(daily.index.min(), daily.index.max())
        daily = daily.reindex(idx, fill_value=0).reset_index().set_index('index')
        # 賺賠次數&金額
        daily['win'] = (daily['數量'] * (daily['勝率'] / 100))
        daily['win'] = daily['win'].apply(lambda x: round(x, 0))
        daily['loss'] = daily['數量'] - daily['win']

        # 原始績效
        origin_profit = int(daily['獲利(¤)'].sum())
        origin_mdd = int(equity(daily)['DD'].max())
        origin_mdd_pct = equity(daily)['DD_pct'].max()
        detail = daily

        for fq in freq:
            # 期間獲利
            result = pd.DataFrame(detail['獲利(¤)'].rolling(fq, min_periods=1).sum().shift(1)).rename(
                columns={'獲利(¤)': 'pnl'})
            result['獲利(¤)'] = detail['獲利(¤)']
            result['max_pnl'] = result['pnl'].rolling(fq, min_periods=1).max()
            result['DD'] = result['max_pnl'] - result['pnl']
            result['MDD'] = result['DD'].rolling(fq, min_periods=1).max()
            result['DD_pct'] = result['DD'] / result['max_pnl']
            result['MDD_pct'] = result['DD_pct'].rolling(fq, min_periods=1).max()
            # 平均獲利
            result['mean'] = detail['獲利(¤)'].rolling(fq, min_periods=1).mean().shift(1).round(2)
            # result['mean(%)'] = detail['獲利(%)'].rolling(fq, min_periods=1).mean().shift(1)
            # 風報比
            result['hazard'] = result['pnl'] / result['MDD'].apply(lambda x: 1 if not x else x)
            # 波動率
            result['volatility'] = detail['獲利(¤)'].rolling(fq, min_periods=1).std().shift(1)
            # result['volatility(%)'] = detail['獲利(%)'].rolling(fq, min_periods=1).std().shift(1)
            # 賺賠次數
            result['win_count'] = detail['win'].rolling(fq, min_periods=1).sum().shift(1)
            result['loss_count'] = detail['loss'].rolling(fq, min_periods=1).sum().shift(1)
            # 賺賠 avg
            result['pnl_w'] = detail['pnl_w'].rolling(fq, min_periods=1).sum().shift(1)
            result['avg_w'] = (result['pnl_w'] / result['win_count']).round(2)
            result['pnl_l'] = detail['pnl_l'].rolling(fq, min_periods=1).sum().shift(1)
            result['avg_l'] = (result['pnl_l'] / result['loss_count'].apply(lambda x: 1 if not x else x)).round(2)
            # 賺賠比
            result['avg_wl'] = result['avg_w'] / -result['avg_l'].apply(lambda x: -1 if not x else x)
            # 勝率
            result['winrate'] = result['win_count'] / detail['數量'].rolling(fq, min_periods=1).sum().shift(1)
            result['winrate'] = result['winrate'].round(4) * 100
            # 獲利因子
            result['pf_factor'] = result['pnl_w'] / -result['pnl_l'].apply(lambda x: -1 if not x else x)
            # Kelly
            result['kelly'] = ((result['winrate'] / 100) * (result['avg_wl'] + 1) - 1) / result['avg_wl']
            # sharpe
            # result['sharpe'] = (result['mean(%)'] / result['volatility(%)']) * (250 ** 0.5)

            # 原始結果輸出
            path = f'period_report/{cat_name}/{strategy_name}.xlsx'
            with pd.ExcelWriter(path, engine="openpyxl", mode='a') as writer:
                result.to_excel(writer, sheet_name=freq_dict[fq])
            wb = load_workbook(path)
            # 刪掉預設sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            wb.save(path)

            # 篩選因子
            for crit in criteria_dict:
                column = 3 + 2 * freq.index(fq)
                # 篩選標準
                for c in criteria_dict[crit]:
                    # criteria為result中符合條件期間(dtype: bool)
                    if crit == 'last':
                        # 前段期間獲利>=0
                        criteria = result['pnl'].shift(c) >= 0
                    elif type(c) == str:
                        # lambda回傳該點在該點以前資料的百分位數
                        if criteria_operator[crit] == '>':
                            criteria = result[crit].rolling(len(result[crit]), min_periods=1).apply(
                                lambda x: x.rank(pct=True).values[-1]) > int(c.split('%')[0]) / 100
                        else:
                            criteria = result[crit].rolling(len(result[crit]), min_periods=1).apply(
                                lambda x: x.rank(pct=True).values[-1]) < int(c.split('%')[0]) / 100
                    else:
                        if criteria_operator[crit] == '>':
                            criteria = result[crit] > c
                        else:
                            criteria = result[crit] < c

                    # 篩後獲益
                    after_profit = result[criteria]['獲利(¤)'].sum()
                    # 獲益變化(百分比)
                    profit_diff = round((after_profit - origin_profit) * 100 / origin_profit, 2)
                    # 篩後DD
                    after_mdd = equity(result[criteria])['DD'].max()
                    after_mdd = int(after_mdd) if after_mdd == after_mdd else 0
                    # MDD變化百分比(>0為減少)
                    mdd_diff = round((origin_mdd - after_mdd) * 100 / origin_mdd, 2)

                    # 寫入excel
                    wb_all[crit].cell(row=row, column=column, value=profit_diff).border = Border(top=thin, bottom=thin)
                    wb_all[crit].cell(row=row, column=column + 1, value=mdd_diff).border = Border(top=thin, bottom=thin)
                    fill(row, column, profit_diff, wb_all[crit])
                    fill(row, column + 1, mdd_diff, wb_all[crit])

                    # 下一標準
                    column += 6
            for fa in factor:
                if fa not in factor_mask:
                    column = 3 + 2 * freq.index(fq)
                    df = pd.read_excel(f"factor/report/{fa}.xlsx", sheet_name=freq_dict[fq], index_col=0)
                    # 將長度縮為策略報告長度
                    len_mask = (df.index >= result.index.min()) & (df.index <= result.index.max())
                    for ec in econ_criteria_dict:
                        for e in econ_criteria_dict[ec]:
                            if type(e) == str:
                                # lambda回傳該點在該點以前資料的百分位數
                                if criteria_operator[ec] == '>':
                                    criteria = df[ec].shift(1)[len_mask].rolling(len(df[ec]), min_periods=1).apply(
                                        lambda x: x.rank(pct=True).values[-1]) > int(e.split('%')[0]) / 100
                                else:
                                    criteria = df[ec].shift(1)[len_mask].rolling(len(df[ec]), min_periods=1).apply(
                                        lambda x: x.rank(pct=True).values[-1]) < int(e.split('%')[0]) / 100
                            else:
                                if criteria_operator[ec] == '>':
                                    criteria = df[ec].shift(1)[len_mask] > e
                                else:
                                    criteria = df[ec].shift(1)[len_mask] > e
                            """
                            after_profit: 篩後獲益
                            after_mdd: 篩後DD
                            """
                            if len(result) > len(criteria):
                                after_profit = result[result.index >= criteria.index.min()][criteria]['獲利(¤)'].sum()
                                after_mdd = equity(result[result.index >= criteria.index.min()][criteria])['DD'].max()
                            else:
                                after_profit = result[criteria]['獲利(¤)'].sum()
                                after_mdd = equity(result[criteria])['DD'].max()

                            # 獲益變化(百分比)
                            profit_diff = round((after_profit - origin_profit) * 100 / origin_profit, 2)
                            # MDD變化百分比(>0為減少)
                            after_mdd = int(after_mdd) if after_mdd == after_mdd else 0
                            mdd_diff = round((origin_mdd - after_mdd) * 100 / origin_mdd, 2)
                            # 寫入excel
                            wb_all[fa].cell(row=row, column=column, value=profit_diff).border = Border(top=thin,
                                                                                                       bottom=thin)
                            wb_all[fa].cell(row=row, column=column + 1, value=mdd_diff).border = Border(top=thin,
                                                                                                        bottom=thin)
                            fill(row, column, profit_diff, wb_all[fa])
                            fill(row, column + 1, mdd_diff, wb_all[fa])
                            column += 6
                else:
                    continue
        # 下一策略
        row += 1

    for cr in criteria_dict:
        # 說明欄
        description(row, wb_all[cr])
        # 合併類別欄
        merge_category(category, wb_all[cr])
    for fa in factor:
        if fa not in factor_mask:
            description(row, wb_all[fa])
            merge_category(category, wb_all[fa])
        else:
            continue
    output = 'output.xlsx'
    wb_all.save(output)
