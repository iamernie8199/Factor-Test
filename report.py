import os
from glob import glob

import pandas as pd
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from factor_test import equity
from setting import freq, freq_dict


def get_file():
    """
    :return: 報表檔名, 商品分類
    """
    # 取得所有報表檔名
    fi = glob('StrategyReport-1/*/*/*.xls*')
    # 避免讀到暫存檔
    for _ in glob('StrategyReport-1/*/*/~*.xls*'):
        fi.remove(_)
    ca = list(set(_.split('\\')[1] for _ in fi))
    ca.sort()
    return fi, ca


if __name__ == "__main__":
    # 取得報表檔名&類別
    file, category = get_file()
    for f in tqdm(file):
        # 策略名稱
        strategy_name = f.split(" ")[2].replace(";", "")
        # 策略分類
        cat_name = f.split('\\')[1]

        # 檢查資料夾
        if not os.path.exists(f"period_report/{cat_name}"):
            os.makedirs(f"period_report/{cat_name}")
        # 建立空白xlsx
        wb = Workbook()
        wb.save(f"period_report/{cat_name}/{strategy_name}.xlsx")

        # 切出日報酬(持倉損益)
        trade = pd.read_excel(f, sheet_name='週期性分析')
        end = trade[trade['Daily Period Analysis'].isin(['Daily Rolling Period Analysis'])].index.values[0]
        daily = trade[0:end]
        daily = daily[daily['Daily Period Analysis'].notna()].reset_index(drop=True)
        daily = daily.set_axis(list(daily.iloc[0]), axis=1, inplace=False)[1:].reset_index(drop=True)
        # 確保欄位名稱統一
        daily = daily.rename(columns={daily.columns[1]: '獲利(¤)',
                                      daily.columns[-1]: '勝率',
                                      '毛利': 'pnl_w',
                                      '毛損': 'pnl_l',
                                      '交易次數': '數量'})
        # 轉換為可計算型別
        daily = daily.astype({'獲利(¤)': 'int',
                              '獲利(%)': 'float',
                              '數量': 'int'})
        daily['期間'] = pd.to_datetime(daily['期間'])
        # 以日期為index
        daily = daily.set_index('期間')
        # 添補缺失日期
        idx = pd.date_range(daily.index.min(), daily.index.max())
        daily = daily.reindex(idx, fill_value=0).reset_index().set_index('index')
        # 賺賠次數&金額
        daily['win'] = (daily['數量'] * (daily['勝率'] / 100))
        daily['win'] = daily['win'].apply(lambda x: round(x, 0))
        daily['loss'] = daily['數量'] - daily['win']

        detail = daily

        for fq in freq:
            # 期間獲利
            result = pd.DataFrame(detail['獲利(¤)'].rolling(fq, min_periods=1).sum().shift(1)).rename(
                columns={'獲利(¤)': 'pnl'})
            result['獲利(¤)'] = detail['獲利(¤)']
            result['max_pnl'] = result['pnl'].rolling(fq, min_periods=1).max()
            result['DD'] = result['max_pnl'] - result['pnl']
            result['MDD'] = result['DD'].rolling(fq, min_periods=1).max()
            result['DD_pct'] = result['DD'] / result['max_pnl']
            result['MDD_pct'] = result['DD_pct'].rolling(fq, min_periods=1).max()
            # 平均獲利
            result['mean'] = detail['獲利(¤)'].rolling(fq, min_periods=1).mean().shift(1).round(2)
            # result['mean(%)'] = detail['獲利(%)'].rolling(fq, min_periods=1).mean().shift(1)
            # 風報比
            result['hazard'] = result['pnl'] / result['MDD'].apply(lambda x: 1 if not x else x)
            # 波動率
            result['volatility'] = detail['獲利(¤)'].rolling(fq, min_periods=1).std().shift(1)
            # result['volatility(%)'] = detail['獲利(%)'].rolling(fq, min_periods=1).std().shift(1)
            # 賺賠次數
            result['win_count'] = detail['win'].rolling(fq, min_periods=1).sum().shift(1)
            result['loss_count'] = detail['loss'].rolling(fq, min_periods=1).sum().shift(1)
            # 賺賠 avg
            result['pnl_w'] = detail['pnl_w'].rolling(fq, min_periods=1).sum().shift(1)
            result['avg_w'] = (result['pnl_w'] / result['win_count']).round(2)
            result['pnl_l'] = detail['pnl_l'].rolling(fq, min_periods=1).sum().shift(1)
            result['avg_l'] = (result['pnl_l'] / result['loss_count'].apply(lambda x: 1 if not x else x)).round(2)
            # 賺賠比
            result['avg_wl'] = result['avg_w'] / -result['avg_l'].apply(lambda x: -1 if not x else x)
            # 勝率
            result['winrate'] = result['win_count'] / detail['數量'].rolling(fq, min_periods=1).sum().shift(1)
            result['winrate'] = result['winrate'].round(4) * 100
            # 獲利因子
            result['pf_factor'] = result['pnl_w'] / -result['pnl_l'].apply(lambda x: -1 if not x else x)
            # Kelly
            result['kelly'] = ((result['winrate'] / 100) * (result['avg_wl'] + 1) - 1) / result['avg_wl']
            # sharpe
            # result['sharpe'] = (result['mean(%)'] / result['volatility(%)']) * (250 ** 0.5)

            # 原始結果輸出
            path = f'period_report/{cat_name}/{strategy_name}.xlsx'
            with pd.ExcelWriter(path, engine="openpyxl", mode='a') as writer:
                result.to_excel(writer, sheet_name=freq_dict[fq])
            wb = load_workbook(path)
            # 刪掉預設sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            wb.save(path)
