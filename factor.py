import os
from glob import glob

import numpy as np
import pandas as pd
import statsmodels.tsa.api as smt
from matplotlib import pyplot as plt
from openpyxl import Workbook, load_workbook
from scipy.stats import probplot
from seaborn import histplot, lineplot
from statsmodels.api import qqplot
from tqdm import tqdm

from preprocess import yahoo, dq2, bls, eia
from setting import freq, freq_dict, longterm


def hurst(ts=None, lags=None):
    """
    USAGE:
        Returns the Hurst Exponent of the time series
        hurst < 0.5: mean revert
        hurst = 0.5: random
        hurst > 0.5: trend

    PARAMETERS:
        ts[,]   a time-series, with 100+ elements
                ( or [ None, ] that produces a demo run )

    RETURNS:
        float - a Hurst Exponent approximation
    """
    if ts is None:
        ts = [None, ]
    if lags is None:
        lags = [2, 100]
    if ts[0] is None:
        """
        DEMO: Create a SYNTH Geometric Brownian Motion, Mean-Reverting and Trending Series
        """
        # a Geometric Brownian Motion[log(1000 + rand), log(1000 + rand + rand ),... log(  1000 + rand + ... )]
        gbm = np.log(1000 + np.cumsum(np.random.randn(100000)))
        # a Mean-Reverting Series    [log(1000 + rand), log(1000 + rand        ),... log(  1000 + rand       )]
        mr = np.log(1000 + np.random.randn(100000))
        # a Trending Series          [log(1001 + rand), log(1002 + rand + rand ),... log(101000 + rand + ... )]
        tr = np.log(1000 + np.cumsum(1 + np.random.randn(100000)))

        # Output the Hurst Exponent for each of the above SYNTH series
        print("hurst(Geometric Browian Motion): {0: > 12.8f}".format(hurst(gbm)))
        print("hurst(   Mean-Reverting Series): {0: > 12.8f}".format(hurst(mr)))
        print("hurst(         Trending Series): {0: > 12.8f}".format(hurst(tr)))
        return "SYNTH series demo (on hurst())"
    if isinstance(ts, pd.Series):
        ts = ts.dropna().to_list()

    too_short_list = lags[1] + 1 - len(ts)
    if 0 < too_short_list:  # IF NOT:
        # 序列長度不足則以第一筆補滿
        ts = too_short_list * ts[:1] + ts  # PRE-PEND SUFFICIENT NUMBER of [ts[0],]-as-list REPLICAS TO THE LIST-HEAD
    # Create the range of lag values
    lags = range(lags[0], lags[1])
    # Calculate the array of the variances of the lagged differences
    tau = [np.sqrt(np.std(np.subtract(ts[lag:], ts[:-lag]))) for lag in lags]
    # Return the Hurst exponent from the polyfit output ( a linear fit to estimate the Hurst Exponent
    return 2.0 * np.polyfit(np.log(lags), np.log(tau), 1)[0]


def tsplot(y, lags=None, figsize=(10, 8), style='seaborn', title=''):
    """
    時間序列統計圖組

    :param y: 時間序列
    :param lags: lags(可略)
    :param figsize: 圖片設定(可略)
    :param style: 圖片風格(可略)
    :param title: 圖片名稱(可略)
    """
    if not isinstance(y, pd.Series):
        y = pd.Series(y)
    with plt.style.context(style):
        plt.figure(figsize=figsize)
        layout = (4, 2)
        ts_ax = plt.subplot2grid(layout, (0, 0), colspan=2)
        acf_ax = plt.subplot2grid(layout, (1, 0))
        pacf_ax = plt.subplot2grid(layout, (1, 1))
        qq_ax = plt.subplot2grid(layout, (2, 0))
        pp_ax = plt.subplot2grid(layout, (2, 1))
        hist = plt.subplot2grid(layout, (3, 0), colspan=2)

        if y.dtype == complex:
            y = y.apply(lambda x: x.real)
        lineplot(label="Daily", ax=ts_ax, data=y)
        lineplot(label="30avg", ax=ts_ax, data=y.rolling(30, min_periods=2).mean())
        lineplot(label="90avg", ax=ts_ax, data=y.rolling(90, min_periods=2).mean())
        """y.plot(ax=ts_ax)
        y.rolling()"""
        ts_ax.set_title(title + ' TSA')
        smt.graphics.plot_acf(y, lags=lags, ax=acf_ax, alpha=0.5, fft=True)
        smt.graphics.plot_pacf(y, lags=lags, ax=pacf_ax, alpha=0.5, method='ols')
        qqplot(y, line='s', ax=qq_ax)
        qq_ax.set_title('QQ Plot')
        probplot(y, sparams=(y.mean(), y.std()), plot=pp_ax)
        histplot(data=y, kde=True, ax=hist)
        hist.set_title('Histogram')
        plt.tight_layout()
        plt.savefig(f'factor/report/pic/{title} TSA.jpg')
        plt.close()


if __name__ == "__main__":
    if not os.path.exists('factor/report'):
        os.makedirs('factor/report')
    if not os.path.exists('factor/report/pic'):
        os.makedirs('factor/report/pic')
    if not os.path.exists('factor/clean/'):
        os.makedirs('factor/clean/')

    bls()
    eia()
    yahoo()
    dq2()

    file = glob('factor/clean/*.csv')
    for f in tqdm(file):
        # 因子代號
        factor_name = f.split('\\')[-1].replace('.csv', '')

        factor = pd.read_csv(f)
        if factor_name not in longterm:
            # 建立空白xlsx
            path = f'factor/report/{factor_name}.xlsx'
            wb = Workbook()
            wb.save(path)

            factor['Date'] = pd.to_datetime(factor['Date'])
            # 將日期設為index
            factor = factor.set_index('Date').sort_index()
            # 轉換格式
            factor = factor.astype('float')
            # 簡單收益率
            factor['pct_change'] = factor['Close'].pct_change()
            # 處理輕原油負值
            if (factor['Close'] < 0).any():
                factor['log_ret'] = np.log(factor['Close'].apply(lambda x: complex(x))) - np.log(
                    factor['Close'].shift(1).apply(lambda x: complex(x)))
            else:
                factor['log_ret'] = np.log(factor['Close']) - np.log(factor['Close'].shift(1))
            factor['range'] = factor['High'] - factor['Low']
            factor['Close[1]'] = factor['Close'].shift(1)
            factor['True Range'] = factor[['High', 'Close[1]']].max(axis=1) - factor[['Low', 'Close[1]']].min(axis=1)
            # 時序統計圖
            tsplot(factor.dropna()['Close'], title=factor_name + ' Close')
            tsplot(factor.dropna()['log_ret'], title=factor_name + ' log return')
            tsplot(factor['range'], title=factor_name + ' range')
            tsplot(factor['True Range'], title=factor_name + ' TR')
            # 補上缺少日期
            idx = pd.date_range(factor.index.min(), factor.index.max())
            # 缺失值以上一筆資料填補
            factor = factor.reindex(idx, method="ffill").reset_index().set_index('index')
            # 重新計算
            factor['pct_change'] = factor['Close'].pct_change()
            # 未交易日期: o=h=l=c
            factor['High'] = factor.apply(lambda x: x['Close'] if x['pct_change'] == 0 else x['High'], axis=1)
            factor['Low'] = factor.apply(lambda x: x['Close'] if x['pct_change'] == 0 else x['Low'], axis=1)
            factor['Open'] = factor.apply(lambda x: x['Close'] if x['pct_change'] == 0 else x['Open'], axis=1)
            factor['Close[1]'] = factor['Close'].shift(1)
            factor['range'] = factor['High'] - factor['Low']
            factor['True Range'] = factor[['High', 'Close[1]']].max(axis=1) - factor[['Low', 'Close[1]']].min(axis=1)
            if (factor['Close'] < 0).any():
                factor['log_ret'] = np.log(factor['Close'].apply(lambda x: complex(x))) - np.log(
                    factor['Close'].shift(1).apply(lambda x: complex(x)))
            else:
                factor['log_ret'] = np.log(factor['Close']) - np.log(factor['Close'].shift(1))
            factor = factor.dropna()

            for fq in freq:
                factor_result = pd.DataFrame(factor['log_ret'].rolling(fq, min_periods=1).sum())
                factor_result['return'] = np.exp(factor_result['log_ret'])
                # 若有虛部則捨去
                factor_result['log_ret'] = np.real(factor['log_ret'])
                factor_result['abs_r'] = abs(factor_result['return'] - 1)
                factor_result['High'] = factor['High'].rolling(fq, min_periods=1).max()
                factor_result['Low'] = factor['Low'].rolling(fq, min_periods=1).min()
                factor_result['range'] = factor_result['High'] - factor_result['Low']
                factor_result['ATR'] = factor['True Range'].ewm(span=fq, min_periods=1, adjust=False).mean()
                # rolling Hurst exponent
                # factor_result['Hurst_exponent'] = factor['Close'].rolling(fq, min_periods=1).apply(lambda x: hurst(x, [2, int(fq/2)]))
                """
                # 若有成交量則計算
                try:
                    factor_result['Vol'] = factor['Volume'].rolling(fq, min_periods=1).sum()
                    # factor_result['Vol_diff'] = factor_result['Vol'].diff(1)
                except:
                    pass
                # 清除全0項
                factor_result = factor_result.loc[:, (factor_result != 0).any(axis=0)]
                # 刪除區間最高/低價
                factor_result = factor_result.drop(columns=['High', 'Low'])
                """
                with pd.ExcelWriter(path, engine="openpyxl", mode='a') as writer:
                    factor_result.to_excel(writer, sheet_name=freq_dict[fq])
            wb = load_workbook(path)
            # 刪掉預設sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            wb.save(path)
        else:
            continue
