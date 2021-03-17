from glob import glob

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from tqdm import tqdm

from setting import freq, freq_dict, freq_dict_zh, criteria_dict, criteria_operator, econ_criteria_dict, factor_mask


def equity(data):
    """
    新增權益曲線/高點&DD欄位
    """
    dataframe = data.copy()
    dataframe['權益'] = dataframe['獲利(¤)'].sort_index().cumsum()
    dataframe['權益高點'] = dataframe['權益'].sort_index().cummax()
    dataframe['DD'] = dataframe['權益高點'] - dataframe['權益']
    dataframe['DD_pct'] = dataframe['DD'] / dataframe['權益高點'].apply(lambda x: 1 if x <= 0 else x)
    dataframe['DD_pct'] = dataframe['DD_pct']
    return dataframe


def freq_title(sheet, col):
    """
    新增header
    :param sheet: openpyxl's ws
    :param col: current column
    :return: updated current column
    """
    init_color = 13434879  # 0xccffff
    for _ in freq_dict_zh:
        sheet.cell(row=3, column=col, value=freq_dict_zh[_]).alignment = Alignment(horizontal="center",
                                                                                   vertical="center")
        sheet.cell(row=3, column=col).fill = PatternFill("solid", fgColor=hex(init_color).split('x')[-1])
        sheet.cell(row=4, column=col, value='績效').alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=4, column=col + 1, value='MDD').alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=4, column=col + 2, value='風報比').alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=4, column=col).border = Border(left=thin)
        sheet.cell(row=4, column=col + 2).border = Border(right=thin)
        sheet.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
        sheet.cell(row=3, column=col).border = Border(left=thin, right=thin)
        col += 3
        init_color -= 3355392
    return col


def fill(r, col, item, sheet):
    """
    :param r: row
    :param col: col
    :param item: Pnl/MDD
    :param sheet: openpyxl's ws
    """
    if item == 100:  # for (MDD==0), 但可能有例外
        color = "808080"
    elif item > 50:
        color = "FF6600"
    elif item > 25:
        color = "FFCC00"
    elif item > 0:
        color = "C0C0C0"
    else:
        color = "808080"
    sheet.cell(row=r, column=col).fill = PatternFill("solid", fgColor=color)


def description(r, sheet):
    """
    :param r: row(工作表最底部)
    :param sheet: openpyxl's ws
    :return:
    """
    sheet.cell(row=r, column=2, value='>50').fill = PatternFill("solid", fgColor='FF6600')
    sheet.cell(row=r, column=2).border = Border(top=thin, left=thin, right=thin)
    sheet.cell(row=r + 1, column=2, value='>25').fill = PatternFill("solid", fgColor='FFCC00')
    sheet.cell(row=r + 1, column=2).border = Border(left=thin, right=thin)
    sheet.cell(row=r + 2, column=2, value='>0').fill = PatternFill("solid", fgColor='C0C0C0')
    sheet.cell(row=r + 2, column=2).border = Border(left=thin, right=thin)
    sheet.cell(row=r + 3, column=2, value='<=0').fill = PatternFill("solid", fgColor='808080')
    sheet.cell(row=r + 3, column=2).border = Border(left=thin, right=thin, bottom=thin)


def merge_category(cat, sheet):
    r = 5
    for ca in cat:
        ca_len = len(glob(f'period_report/{ca}/*.xlsx'))
        # 合併儲存格
        sheet.merge_cells(start_row=r, start_column=1, end_row=r + ca_len - 1, end_column=1)
        sheet.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        r += ca_len


def get_file():
    """
    :return: 報表檔名, 商品分類
    """
    # 取得所有報表檔名
    fi = glob('period_report/*/*.xlsx')
    # 避免讀到暫存檔
    for _ in glob('period_report/*/~*.xls*'):
        fi.remove(_)
    ca = list(set(_.split('\\')[1] for _ in fi))
    ca.sort()
    return fi, ca


def get_factor():
    """
    :return: factor list
    """
    fac = glob('factor/report/*.xlsx')
    fac = [_.split('\\')[-1].replace(".xlsx", "") for _ in fac]
    return fac


def strategy_title(sheet):
    """
    策略名稱欄位
    :param sheet: openpyxl's ws
    """
    sheet.cell(row=4, column=2).border = Border(bottom=thin)
    sheet.cell(row=4, column=1).border = Border(bottom=thin)
    sheet.merge_cells(start_row=1, start_column=1, end_row=4, end_column=2)
    sheet.column_dimensions["B"].bestFit = True


def strategy_n_category(workbook, sheet, ca, stra, r):
    """
    寫入策略名稱/類別
    :param workbook: openpyxl's workbook
    :param sheet: sheet name
    :param ca: category
    :param stra: strategy name
    :param r: row
    """
    workbook[sheet].cell(row=r, column=1, value=ca)
    workbook[sheet].cell(row=r, column=2, value=stra)
    workbook[sheet].cell(row=r, column=2).border = Border(right=thin)


if __name__ == "__main__":
    # excel框線
    thin = Side(border_style="thin", color="000000")

    # 取得報表檔名&類別
    file, category = get_file()
    factor = get_factor()
    # 生成篩後結果excel
    wb_all = Workbook()
    ws = wb_all.active

    for cr in criteria_dict:
        # 建立工作表
        wb_all.create_sheet(cr)
        # 策略名稱欄位
        strategy_title(wb_all[cr])
        # 績效Factor
        wb_all[cr].cell(row=1, column=3, value=cr)
        space = 3 * 3 * len(criteria_dict[cr])
        wb_all[cr].merge_cells(start_row=1, start_column=3, end_row=1, end_column=3 + space - 1)
        wb_all[cr].cell(row=1, column=3).alignment = Alignment(horizontal="center", vertical="center")
        wb_all[cr].cell(row=1, column=3).border = Border(left=thin, right=thin)
        column = 3
        # Factor's criteria
        for i in criteria_dict[cr]:
            wb_all[cr].cell(row=2, column=column, value=criteria_operator[cr] + str(i))
            wb_all[cr].cell(row=2, column=column).alignment = Alignment(horizontal="center", vertical="center")
            wb_all[cr].merge_cells(start_row=2, start_column=column, end_row=2, end_column=column + 8)
            wb_all[cr].cell(row=2, column=column).border = Border(left=thin, right=thin)
            # freq
            column = freq_title(wb_all[cr], column)

    factor_criteria_num = 0
    for _ in econ_criteria_dict:
        factor_criteria_num += len(econ_criteria_dict[_])

    for _ in factor:
        if _ not in factor_mask:
            wb_all.create_sheet(_)
            # 策略名稱欄位
            strategy_title(wb_all[_])
            column = 3
            for ec in econ_criteria_dict:
                wb_all[_].cell(row=1, column=column, value=ec)
                space = 3 * 3 * len(econ_criteria_dict[ec])
                wb_all[_].merge_cells(start_row=1, start_column=column, end_row=1, end_column=column + space - 1)
                wb_all[_].cell(row=1, column=column).alignment = Alignment(horizontal="center", vertical="center")
                wb_all[_].cell(row=1, column=column).border = Border(left=thin, right=thin)
                for i in econ_criteria_dict[ec]:
                    wb_all[_].cell(row=2, column=column, value=criteria_operator[ec] + str(i))
                    wb_all[_].cell(row=2, column=column).alignment = Alignment(horizontal="center", vertical="center")
                    wb_all[_].merge_cells(start_row=2, start_column=column, end_row=2, end_column=column + 8)
                    wb_all[_].cell(row=2, column=column).border = Border(left=thin, right=thin)
                    # freq
                    column = freq_title(wb_all[_], column)
        else:
            continue
    # 刪除預設空白工作表
    wb_all.remove(ws)

    row = 5
    for f in tqdm(file):
        column = 3
        # 策略名稱
        strategy_name = f.split("\\")[-1].replace(".xlsx", "")
        # 策略分類
        cat_name = f.split('\\')[1]

        for cr in criteria_dict:
            strategy_n_category(wb_all, cr, cat_name, strategy_name, row)
        for fa in factor:
            if fa not in factor_mask:
                strategy_n_category(wb_all, fa, cat_name, strategy_name, row)
            else:
                continue

        for fq in freq:
            result = pd.read_excel(f, sheet_name=freq_dict[fq], index_col=0)
            origin_profit = int(result['獲利(¤)'].sum())
            tmp = equity(result)
            origin_mdd = int(tmp['DD'].max())
            origin_mdd_pct = tmp['DD_pct'].max()
            origin_hazard = origin_profit / origin_mdd
            # 篩選因子
            for crit in criteria_dict:
                column = 3 + 3 * freq.index(fq)
                # 篩選標準
                for c in criteria_dict[crit]:
                    # criteria為result中符合條件期間(dtype: bool)
                    if crit == 'last':
                        # 前段期間獲利>=0
                        criteria = result['pnl'].shift(c) >= 0
                    elif type(c) == str:
                        # lambda回傳該點在該點以前資料的百分位數
                        if criteria_operator[crit] == '>':
                            criteria = result[crit].rolling(len(result[crit]), min_periods=1).apply(
                                lambda x: x.rank(pct=True).values[-1]) > int(c.split('%')[0]) / 100
                        else:
                            criteria = result[crit].rolling(len(result[crit]), min_periods=1).apply(
                                lambda x: x.rank(pct=True).values[-1]) < int(c.split('%')[0]) / 100
                    else:
                        if criteria_operator[crit] == '>':
                            criteria = result[crit] > c
                        else:
                            criteria = result[crit] < c

                    # 篩後獲益
                    after_profit = result[criteria]['獲利(¤)'].sum()
                    # 獲益變化(百分比)
                    profit_diff = round((after_profit - origin_profit) * 100 / origin_profit, 2)
                    # 篩後DD
                    after_mdd = equity(result[criteria])['DD'].max()
                    after_mdd = int(after_mdd) if after_mdd == after_mdd else 0
                    # MDD變化百分比(>0為減少)
                    mdd_diff = round((origin_mdd - after_mdd) * 100 / origin_mdd, 2)
                    #  風暴比
                    after_hazard = after_profit / (after_mdd if after_mdd else 1)
                    hazard_diff = round((after_hazard - origin_hazard) * 100 / origin_hazard, 2)
                    # 寫入excel
                    wb_all[crit].cell(row=row, column=column, value=profit_diff).border = Border(top=thin, bottom=thin)
                    wb_all[crit].cell(row=row, column=column + 1, value=mdd_diff).border = Border(top=thin, bottom=thin)
                    wb_all[crit].cell(row=row, column=column + 2, value=hazard_diff).border = Border(top=thin,
                                                                                                     bottom=thin)
                    fill(row, column, profit_diff, wb_all[crit])
                    fill(row, column + 1, mdd_diff, wb_all[crit])
                    fill(row, column + 2, hazard_diff, wb_all[crit])

                    # 下一標準
                    column += 9
            for fa in factor:
                if fa not in factor_mask:
                    column = 3 + 3 * freq.index(fq)
                    df = pd.read_excel(f"factor/report/{fa}.xlsx", sheet_name=freq_dict[fq], index_col=0)
                    # 將長度縮為策略報告長度
                    len_mask = (df.index >= result.index.min()) & (df.index <= result.index.max())
                    for ec in econ_criteria_dict:
                        for e in econ_criteria_dict[ec]:
                            if type(e) == str:
                                # lambda回傳該點在該點以前資料的百分位數
                                if criteria_operator[ec] == '>':
                                    criteria = df[ec].shift(1)[len_mask].rolling(len(df[ec]), min_periods=1).apply(
                                        lambda x: x.rank(pct=True).values[-1]) > int(e.split('%')[0]) / 100
                                else:
                                    criteria = df[ec].shift(1)[len_mask].rolling(len(df[ec]), min_periods=1).apply(
                                        lambda x: x.rank(pct=True).values[-1]) < int(e.split('%')[0]) / 100
                            else:
                                if criteria_operator[ec] == '>':
                                    criteria = df[ec].shift(1)[len_mask] > e
                                else:
                                    criteria = df[ec].shift(1)[len_mask] > e
                            """
                            after_profit: 篩後獲益
                            after_mdd: 篩後DD
                            """
                            if len(result) > len(criteria):
                                after_profit = result[result.index >= criteria.index.min()][criteria]['獲利(¤)'].sum()
                                after_mdd = equity(result[result.index >= criteria.index.min()][criteria])['DD'].max()
                            else:
                                after_profit = result[criteria]['獲利(¤)'].sum()
                                after_mdd = equity(result[criteria])['DD'].max()

                            # 獲益變化(百分比)
                            profit_diff = round((after_profit - origin_profit) * 100 / origin_profit, 2)
                            # MDD變化百分比(>0為減少)
                            after_mdd = int(after_mdd) if after_mdd == after_mdd else 0
                            mdd_diff = round((origin_mdd - after_mdd) * 100 / origin_mdd, 2)
                            # 風報比
                            after_hazard = after_profit / (after_mdd if after_mdd else 1)
                            hazard_diff = round((after_hazard - origin_hazard) * 100 / origin_hazard, 2)
                            # 寫入excel
                            wb_all[fa].cell(row=row, column=column, value=profit_diff).border = Border(top=thin,
                                                                                                       bottom=thin)
                            wb_all[fa].cell(row=row, column=column + 1, value=mdd_diff).border = Border(top=thin,
                                                                                                        bottom=thin)
                            wb_all[fa].cell(row=row, column=column + 2, value=hazard_diff).border = Border(top=thin,
                                                                                                        bottom=thin)
                            fill(row, column, profit_diff, wb_all[fa])
                            fill(row, column + 1, mdd_diff, wb_all[fa])
                            fill(row, column + 2, hazard_diff, wb_all[fa])
                            column += 9
                else:
                    continue
        # 下一策略
        row += 1
    for d in dict(criteria_dict, **{f: [] for f in factor}):
        # 說明欄
        description(row, wb_all[d])
        # 合併類別欄
        merge_category(category, wb_all[d])
    output = 'output.xlsx'
    wb_all.save(output)
